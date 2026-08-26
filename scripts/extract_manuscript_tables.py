#!/usr/bin/env python3
"""Copy the submitted supplementary tables into the site's versioned data dir.

Run by hand, never by a build. The site build reads only
GPCompaReports_v2/data/manuscript_<date>/, so it never depends on the source
workbook being present, and the source lives on a removable volume.

    python3 scripts/extract_manuscript_tables.py \
        --source "/media/yamir/OS/Users/yamam/Downloads/GPCompaRe__Supplementary_Tables.xlsx"

The source workbook is opened read-only and is never written to. Figure 3
values are transcribed by hand into figure3_enrichment.json rather than parsed
from a PDF, so changing them is a deliberate, reviewable edit.
"""
import argparse
import csv
import hashlib
import pathlib
import sys

ROOT = pathlib.Path(__file__).resolve().parent.parent
DEST = ROOT / 'GPCompaReports_v2' / 'data' / 'manuscript_2026-08-24'

SHEETS = [
    ('S3_recurrent_CFR_positions', 'S3_top50_recurrent_cfr_positions.csv', 50),
    ('S4_recurrent_positions_all', 'S4_all_recurrent_positions.csv', 368),
    ('S5_recurrent_CFR_pairs', 'S5_top20_recurrent_cfr_pairs.csv', 20),
]


def header_row(rows):
    """Sheets carry a title and one or more notes above the real header."""
    for i, r in enumerate(rows):
        cells = [('' if c is None else str(c)).strip() for c in r]
        if sum(1 for c in cells if c) >= 4 and cells[0] and ' ' not in cells[0]:
            return i
    raise SystemExit('could not find a header row')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--source', required=True,
                    help='path to GPCompaRe__Supplementary_Tables.xlsx')
    ap.add_argument('--force', action='store_true',
                    help='overwrite the existing extraction')
    args = ap.parse_args()

    src = pathlib.Path(args.source)
    if not src.is_file():
        raise SystemExit('source workbook not found: %s' % src)
    if DEST.exists() and any(DEST.glob('*.csv')) and not args.force:
        raise SystemExit(
            'refusing to overwrite %s without --force. These files are the '
            'scientific authority for the statistics page; replacing them '
            'changes published numbers and needs review.' % DEST)

    import openpyxl
    print('source     : %s' % src)
    print('sha256     : %s' % hashlib.sha256(src.read_bytes()).hexdigest())
    DEST.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    for sheet, out_name, expect in SHEETS:
        if sheet not in wb.sheetnames:
            raise SystemExit('sheet %s missing from the workbook' % sheet)
        rows = list(wb[sheet].iter_rows(values_only=True))
        h = header_row(rows)
        header = [c for c in
                  [('' if c is None else str(c)).strip() for c in rows[h]] if c]
        body = []
        for r in rows[h + 1:]:
            cells = ['' if c is None else str(c) for c in r][:len(header)]
            if any(x.strip() for x in cells):
                body.append(cells)
        if len(body) != expect:
            raise SystemExit(
                '%s: expected %d data rows, found %d. The workbook is not the '
                'one this extraction was written for; check before proceeding.'
                % (sheet, expect, len(body)))
        with (DEST / out_name).open('w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(body)
        print('  wrote %-44s %3d rows' % (out_name, len(body)))
    wb.close()
    print('\nfigure3_enrichment.json is NOT regenerated here. It is transcribed '
          'by hand from the main-figures PDF and the Results paragraph.')
    print('Update PROVENANCE.md if the source workbook changed.')


if __name__ == '__main__':
    sys.exit(main())
